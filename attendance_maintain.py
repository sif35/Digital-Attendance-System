import xlsxwriter as xl
import datetime
import os
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font
import main_names
from attendance_percentage import Percentage


def create_new_attendance_sheet(file_path=None, file_name=None):

    if file_path is None:
        file_path = "./Attendance"

    if os.path.isdir(file_path):
        number_of_file = len(os.listdir("Attendance"))
    else:
        os.mkdir(file_path)
        number_of_file = len(os.listdir("Attendance"))

    if file_name is None:
        attendance = xl.Workbook(f'{file_path}/Attendance Sheet - {number_of_file + 1}.xlsx')
    else:
        attendance = xl.Workbook(f'{file_path}/{file_name}.xlsx')
    attendance_sheet = attendance.add_worksheet("Attendance")

    title_styles = attendance.add_format({'align': 'center',
                                          'bold': True,
                                          'font_size': 18})  # Font size = 18
    attendance_sheet.set_column(0, 0, 40)  # Student Names column width
    attendance_sheet.set_column(1, 1, 15)  # Roll No. column width

    attendance_sheet.write(0, 0, "Student Names", title_styles)
    attendance_sheet.write(0, 1, "Roll No.", title_styles)

    row = 1
    col = 0
    write_format = attendance.add_format({'align': 'center',
                                          'font_size': 14})  # Font size = 14
    for name, roll in main_names.name_dict.values():
        attendance_sheet.write(row, col, name, write_format)
        attendance_sheet.write(row, col + 1, roll, write_format)
        row += 1

    attendance.close()


class Attendance:

    time = datetime.datetime.now()

    def __init__(self, attendance_file_path):

        self.file_path = attendance_file_path

    percentage_file_path = "Attendance/Attendance Percentage/Attendance Sheet Percentage.xlsx"

    def modify_attendance_sheet(self, name_list=None):

        list_of_name_dict = list(main_names.name_dict)
        workbook = openpyxl.load_workbook(self.file_path)  # file_path variable will be use here
        active_sheet = workbook.active

        # row = active_sheet.max_row
        col = active_sheet.max_column

        # active_sheet.merge_cells(start_row=1, start_column=col+1, )
        active_sheet.column_dimensions[get_column_letter(col + 1)].width = 20
        date_cell = active_sheet.cell(row=1, column=col + 1, value=self.time.strftime("%x"))
        date_cell.alignment = Alignment(vertical='center', horizontal='center')
        date_cell.font = Font(bold=True, size=20)

        for name in list_of_name_dict:
            row_for_iteration = list_of_name_dict.index(name) + 2
            col_for_iteration = col + 1
            if name in name_list:
                cell = active_sheet.cell(row=row_for_iteration, column=col_for_iteration,
                                         value='P')  # Print "P" if present
                cell.alignment = Alignment(vertical='center', horizontal='center')
                cell.font = Font(size=14)  # Font size = 14
            else:
                cell = active_sheet.cell(row=row_for_iteration, column=col_for_iteration,
                                         value='A')  # Print "A" if absent
                cell.alignment = Alignment(vertical='center', horizontal='center')
                cell.font = Font(size=14)  # Font size = 14

        workbook.save(self.file_path)  # file_path variable will be use here

    def calculate_number_of_attendance(self):

        attendance_file = openpyxl.load_workbook(self.file_path)

        if os.path.isfile(self.percentage_file_path) is False:
            percentage = Percentage()
            percentage.create_new_percentage_sheet()

        percentage_file = openpyxl.load_workbook(self.percentage_file_path)

        file_sheet = attendance_file.active
        percentage_file_sheet = percentage_file.active

        start_row = 2
        start_col = 3
        max_row = file_sheet.max_row
        max_col = file_sheet.max_column

        student_presents = {}
        nickname_list = []
        fullname_list = []
        number_of_days = max_col - 2  # Number of days stored in the sheet

        # Getting the list of nicknames and corresponding full names
        for nickname, fullname in main_names.name_dict.items():
            nickname_list.append(nickname)
            fullname_list.append(fullname[0])
        print(fullname_list)

        # Creating a dictionary to update students presents
        for name in main_names.name_dict.keys():
            student_presents.update({name: 0})

        # Checking the attendance sheet and calculating absence and presence per student
        for row in range(start_row, max_row + 1):

            sheet_name = file_sheet.cell(row, 1).value
            student_name = nickname_list[fullname_list.index(sheet_name)]

            for col in range(start_col, max_col + 1):

                cell_value = file_sheet.cell(row, col).value
                if cell_value == "P":
                    student_presents[student_name] += 1
                elif cell_value == "A":
                    student_presents[student_name] += 0
                else:
                    continue
        percentage_file_sheet.column_dimensions[get_column_letter(3)].width = 20
        total_present_cell = percentage_file_sheet.cell(1, 3, "Total Present")
        total_present_cell.alignment = Alignment(vertical='center', horizontal='center')
        total_present_cell.font = Font(bold=True, size=14)  # Font size is 14

        percentage_file_sheet.column_dimensions[get_column_letter(4)].width = 20
        total_absent_cell = percentage_file_sheet.cell(1, 4, "Total Absent")
        total_absent_cell.alignment = Alignment(vertical='center', horizontal='center')
        total_absent_cell.font = Font(bold=True, size=14)  # Font size is 14

        for row in range(start_row, max_row+1):

            sheet_name = percentage_file_sheet.cell(row, 1).value
            student_name = nickname_list[fullname_list.index(sheet_name)]

            cell_present = percentage_file_sheet.cell(row, 3, value=student_presents[student_name])
            cell_present.alignment = Alignment(vertical='center', horizontal='center')
            cell_present.font = Font(bold=True, size=14)  # Font size is 14

            cell_absent = percentage_file_sheet.cell(row, 4, value=number_of_days-student_presents[student_name])
            cell_absent.alignment = Alignment(vertical='center', horizontal='center')
            cell_absent.font = Font(bold=True, size=14)  # Font size is 14

        percentage_file.save(self.percentage_file_path)
        attendance_file.close()

# create_new_attendance_sheet()
