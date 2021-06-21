import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font
import main_names
import os
import xlsxwriter as xl


class Percentage:

    file_path = "Attendance/Attendance Percentage"
    file_name = "Attendance Sheet Percentage"

    def create_new_percentage_sheet(self):

        number_of_file = 0

        if os.path.isdir(self.file_path) is False:
            os.mkdir(self.file_path)

        attendance = xl.Workbook(f'{self.file_path}/{self.file_name}.xlsx')
        attendance_sheet = attendance.add_worksheet("Percentage")

        title_styles = attendance.add_format({'align': 'center',
                                              'bold': True,
                                              'font_size': 14})  # Font size = 14
        attendance_sheet.set_column(0, 0, 40)  # Student Names column width
        attendance_sheet.set_column(1, 1, 15)  # Roll No. column width

        attendance_sheet.write(0, 0, "Student Names", title_styles)
        attendance_sheet.write(0, 1, "Roll No.", title_styles)

        row = 1
        col = 0
        write_format = attendance.add_format({'align': 'center',
                                              'font_size': 10})  # Font size = 10
        for name, roll in main_names.name_dict.values():
            attendance_sheet.write(row, col, name, write_format)
            attendance_sheet.write(row, col + 1, roll, write_format)
            row += 1

        attendance.close()

    def calculate_percentage_and_update(self, attendance_file_path):

        if os.path.isfile(f'{self.file_path}/{self.file_name}.xlsx') is False:
            self.create_new_percentage_sheet()

        attendance_file = openpyxl.load_workbook(attendance_file_path)
        percentage_file = openpyxl.load_workbook(f'{self.file_path}/{self.file_name}.xlsx')

        file_sheet = attendance_file.active
        percentage_file_sheet = percentage_file.active

        start_row = 2
        start_col = 3
        max_row = file_sheet.max_row
        max_col = file_sheet.max_column

        student_presents = {}
        nickname_list = []
        fullname_list = []
        number_of_days = max_col - 2  # Number of days stored in the sheet

        # Getting the list of nicknames and corresponding full names
        for nickname, fullname in main_names.name_dict.items():
            nickname_list.append(nickname)
            fullname_list.append(fullname[0])

        # Creating a dictionary to update students presents
        for name in main_names.name_dict.keys():
            student_presents.update({name: 0})

        # Checking the attendance sheet and calculating absence and presence per student
        for row in range(start_row, max_row + 1):

            sheet_name = file_sheet.cell(row, 1).value
            student_name = nickname_list[fullname_list.index(sheet_name)]

            for col in range(start_col, max_col + 1):

                cell_value = file_sheet.cell(row, col).value
                if cell_value == "P":
                    student_presents[student_name] += 1
                elif cell_value == "A":
                    student_presents[student_name] += 0
                else:
                    continue
        # Saving dates
        first_date = file_sheet.cell(1, start_col).value
        last_date = file_sheet.cell(1, max_col).value

        # print(student_presents)
        # print("Number of days: {}".format(number_of_days))

        # Modifying and updating the student_percentage sheet
        max_col_for_percentage = percentage_file_sheet.max_column
        percentage_file_sheet.column_dimensions[get_column_letter(max_col_for_percentage+1)].width = 30

        percentage_title_cell = percentage_file_sheet.cell(row=1, column=max_col_for_percentage+1,
                                                           value="Days: {} - {}".format(first_date, last_date))
        percentage_title_cell.font = Font(bold=True, size=12)  # Font size = 12
        percentage_title_cell.alignment = Alignment(vertical='center', horizontal='center')

        percentage_row = 2

        # Updating the student_percentage
        for name, presence in student_presents.items():
            student_percentage = (presence * 100) / number_of_days
            cell = percentage_file_sheet.cell(row=percentage_row, column=max_col_for_percentage+1,
                                              value=round(student_percentage, 2))
            cell.alignment = Alignment(vertical='center', horizontal='center')
            cell.font = Font(size=14, bold=True)  # Font size = 14
            percentage_row += 1

        percentage_file.save(f'{self.file_path}/{self.file_name}')
        attendance_file.close()
